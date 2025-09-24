#!/usr/bin/env python3
"""
Electron Bridge for OTJ Automation
Accepts JSON input from Electron app and processes it through the OTJ system
"""

import json
import sys
import os
from datetime import datetime

# Import required libraries directly instead of from OTJ_Automation
import warnings
warnings.simplefilter("ignore", UserWarning)

from openpyxl import load_workbook
from datetime import datetime
import time

# Global variables
rowData = []
path = ""

# Required functions copied from OTJ_Automation.py
def wrtTXT(file, line):
    """Write text to a file"""
    try:
        # Ensure directory exists
        log_dir = os.path.dirname(file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)
        
        with open(file, 'a', encoding='utf-8') as filename:
            filename.write(line)
    except Exception as e:
        # If logging fails, print to stderr (which gets captured by Electron)
        print(f"Log write failed: {e}", file=sys.stderr)

def log(line, style):
    """Log function with different styles"""
    currentTime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Use a more reliable log path in the same directory as the script
    log_path = os.path.join(os.path.dirname(__file__), "bridge_log.txt")
    
    if style == 1:
        output = "INFO    - {} - {}\n".format(currentTime, line)
        wrtTXT(log_path, output)
    elif style == 2:
        output = "#ERROR# - {} - {}\n".format(currentTime, line)
        wrtTXT(log_path, output)
    elif style == 999:
        output = "#FATAL# - {} - {}\n".format(currentTime, line)
        wrtTXT(log_path, output)

def findFirstBlankRow():
    """Find the first blank row in the Excel sheet"""
    # Get starting row from settings
    settings = load_settings_from_electron()
    starting_row = settings.get('startingRow', 125)  # Default to 125 if not set
    
    wb = load_workbook(path)
    sheet = wb["OTJ log"]
    
    for row in range(starting_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value is None:
            return row
    
    # If no blank row found, return next row
    return sheet.max_row + 1

def readCellKSB(row, col):
    """Read cell from KSB sheet"""
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]
    
    contents = sheet.cell(row=row, column=col).value
    contents = str(contents) if contents is not None else "None"
    
    if contents == "None":
        return f"NO DATA IN CELL {col}/{row}"
    else:
        return contents

def findInRowKSB(term, style):
    """Find term in KSB sheet row"""
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]
    
    output = []
    total = 0
    
    for iterateCols in range(3, sheet.max_column + 1):
        contents = sheet.cell(row=2, column=iterateCols).value
        contents = str(contents) if contents is not None else ""
        
        if term in contents:
            output.append(f"2/{iterateCols}")
            total += 1
    
    if output and style == 1:
        return output
    elif output and style == 2:
        return total
    else:
        return "Not found"

def findInColKSB(term, col, style):
    """Find term in KSB sheet column"""
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]
    
    output = []
    total = 0
    
    for iterateRows in range(1, sheet.max_row + 1):
        contents = sheet.cell(row=iterateRows, column=col).value
        contents = str(contents) if contents is not None else ""
        
        if term in contents:
            output.append(f"{iterateRows}/{col}")
            total += 1
    
    if output and style == 1:
        return output
    elif output and style == 2:
        return total
    else:
        return "Not found"

def getKSB(module):
    """Get KSB codes for a module"""
    try:
        module = module.upper()
        
        rowsWithX = []
        fullKSB = []
        
        received = findInRowKSB(module, 1)
        if received == "Not found":
            return ""
        
        column = str(received[0])
        column = column.split("/")
        column = int(column[1])
        
        rowsWithXOutput = findInColKSB("x", column, 1)
        if rowsWithXOutput == "Not found":
            return ""
        
        for item in rowsWithXOutput:
            parts = item.split("/")
            number_after_slash = parts[0]
            rowsWithX.append(number_after_slash)
        
        for item in rowsWithX:
            ksb_value = readCellKSB(int(item), 2)
            if ksb_value and not ksb_value.startswith("NO DATA"):
                fullKSB.append(ksb_value[:2])
        
        return str(",".join(fullKSB))
    except Exception as e:
        log(f"getKSB error: {str(e)}", 2)
        return ""

def addDeclaration(declaration_value):
    """Process declaration value and return both declaration and confirmation values"""
    # declaration_value should be 'Yes' or 'No'
    if declaration_value == 'Yes':
        return declaration_value, 'N/A'
    elif declaration_value == 'No':
        # For 'No', confirmation will be provided separately
        return declaration_value, None
    else:
        raise Exception(f"Invalid declaration value: {declaration_value}. Must be 'Yes' or 'No'")

def addConfirmation(confirmation_value):
    """Process confirmation value"""
    # confirmation_value should be 'Yes', 'No', or 'Not applicable'
    valid_options = ['Yes', 'No', 'Not applicable']
    if confirmation_value in valid_options:
        return confirmation_value
    else:
        raise Exception(f"Invalid confirmation value: {confirmation_value}. Must be one of: {', '.join(valid_options)}")

def load_settings_from_electron():
    """Load settings from electron_settings.json file"""
    # Try multiple paths to find the settings file
    script_dir = os.path.dirname(__file__)
    cwd = os.getcwd()
    
    # Possible paths where settings might be located
    # Try user's development location first to preserve existing settings
    dev_project_path = "/Users/andrewjolley/lecture-logger/python"
    
    possible_paths = [
        # User's existing development settings (highest priority)
        dev_project_path,
        # Current development: main python directory
        os.path.join(cwd, 'python'),
        # Development: script directory if running from python/
        script_dir,
        # Built app: relative to working directory
        cwd,
        # Built app: relative to script directory
        os.path.dirname(script_dir),
        # Built app: app.asar structure
        os.path.join(os.path.dirname(cwd), 'python'),
        # Built app: alongside runtime directory
        os.path.join(os.path.dirname(script_dir), '..', 'python'),
        # Built app: in app bundle Contents directory
        os.path.join(cwd, '..', 'python') if 'Contents' in cwd else None,
    ]
    
    # Remove None entries
    possible_paths = [p for p in possible_paths if p is not None]
    
    settings_path = None
    clean_settings_path = None
    
    # Try each possible path until we find the settings file
    for python_dir in possible_paths:
        test_settings_path = os.path.join(python_dir, "electron_settings.json")
        if os.path.exists(test_settings_path):
            settings_path = test_settings_path
            clean_settings_path = os.path.join(python_dir, "electron_settings_clean.json")
            break
    
    # If no existing settings found, use the first path (main python directory)
    if not settings_path:
        python_dir = possible_paths[0]
        settings_path = os.path.join(python_dir, "electron_settings.json")
        clean_settings_path = os.path.join(python_dir, "electron_settings_clean.json")
    
    try:
        with open(settings_path, 'r') as f:
            settings = json.load(f)
        return settings
    except FileNotFoundError:
        # Create settings file from clean template
        try:
            with open(clean_settings_path, 'r') as f:
                clean_settings = json.load(f)
            
            # Create user settings file from clean template
            with open(settings_path, 'w') as f:
                json.dump(clean_settings, f, indent=2)
            
            return clean_settings
        except FileNotFoundError:
            # Fallback if clean template doesn't exist
            default_settings = {
                "excelPath": "",
                "startingRow": 1,
                "verboseLogging": False,
                "timestamp": ""
            }
            with open(settings_path, 'w') as f:
                json.dump(default_settings, f, indent=2)
            return default_settings
    except json.JSONDecodeError:
        return {}  # Return empty dict if JSON is invalid

def get_excel_path():
    """Get Excel file path from Electron settings"""
    global path
    settings = load_settings_from_electron()
    path = settings.get('excelPath', '')
    if not path:
        raise Exception("Excel file path not set in settings. Please configure the Excel file path in the app settings.")
    if not os.path.exists(path):
        raise Exception(f"Excel file not found at path: {path}. Please check the file path in settings.")
    return path

def calculate_academic_year(date_str):
    """Calculate academic year from date string (DD/MM/YYYY)"""
    try:
        parts = date_str.split("/")
        month = int(parts[1])
        year = int(parts[2][-2:])  # Last 2 digits
        
        if month >= 9:
            start_year = year
        else:
            start_year = year - 1
            
        end_year = start_year + 1
        return f"{start_year}/{end_year}"
    except Exception as e:
        raise Exception(f"Failed to calculate academic year: {str(e)}")

def process_form_data(data):
    """Process form data from Electron and write to Excel"""
    global rowData, path
    
    try:
        # Get Excel path
        path = get_excel_path()
        
        # Validate required fields
        required_fields = ['date', 'location', 'activityType', 'moduleCode', 'description', 'details', 'duration']
        for field in required_fields:
            if field not in data or not data[field]:
                raise Exception(f"Missing required field: {field}")
        
        # Validate conditional fields for declaration/confirmation logic
        if 'declaration' in data and data['declaration'] == 'No':
            if 'confirmation' not in data or not data['confirmation']:
                raise Exception("Confirmation field is required when declaration is 'No'")
        
        # Build rowData array in the same order as original script
        rowData = []
        
        # Add date
        date_value = data['date']
        if date_value.lower() == 'today':
            date_value = datetime.now().strftime("%d/%m/%Y")
        rowData.append(date_value)
        
        # Add academic year (calculated from date)
        academic_year = calculate_academic_year(date_value)
        rowData.append(academic_year)
        
        # Add location
        rowData.append(data['location'])
        
        # Add activity type
        rowData.append(data['activityType'])
        
        # Add module code
        rowData.append(data['moduleCode'])
        
        # Add description
        rowData.append(data['description'])
        
        # Add details
        rowData.append(data['details'])
        
        # Add KSB (if module is not "Not applicable")
        if data['moduleCode'] != "Not applicable":
            try:
                ksb = getKSB(data['moduleCode'])
                rowData.append(ksb)
            except Exception as e:
                rowData.append("")  # Empty KSB if error
                log(f"KSB Error: {str(e)}", 2)
        else:
            rowData.append("")  # Empty KSB for "Not applicable"
        
        # Add next steps
        rowData.append(data.get('nextSteps', ''))
        
        # Add duration
        duration = float(data['duration'])
        if duration <= 0 or duration >= 50:
            raise Exception("Duration must be between 0 and 50 hours")
        rowData.append(duration)
        
        # Add declaration and confirmation (new columns M and N)
        log(f"Adding declaration/confirmation - rowData length before: {len(rowData)}", 1)
        declaration_value = data.get('declaration', 'Yes')  # Default to 'Yes' if not provided
        log(f"Declaration value from data: {declaration_value}", 1)
        
        declaration, confirmation = addDeclaration(declaration_value)
        log(f"addDeclaration returned: {declaration}, {confirmation}", 1)
        rowData.append(declaration)
        
        # Add confirmation value
        if confirmation is None:
            # Declaration was 'No', so get confirmation from data
            confirmation_value = data.get('confirmation', 'Not applicable')
            confirmation = addConfirmation(confirmation_value)
            log(f"addConfirmation returned: {confirmation}", 1)
        rowData.append(confirmation)
        log(f"Final rowData length: {len(rowData)}", 1)
        
        # Write to Excel
        wb = load_workbook(path)
        sheet = wb["OTJ log"]
        
        row = findFirstBlankRow()
        
        for col, value in enumerate(rowData, start=3):
            sheet.cell(row=row, column=col, value=value)
        
        wb.save(path)
        
        # Log success
        log(f"Bridge - Successfully wrote data to row {row}", 1)
        
        return {
            "success": True,
            "message": f"Successfully added entry to row {row}",
            "row": row,
            "data": rowData
        }
        
    except Exception as e:
        error_msg = str(e)
        log(f"Bridge Error: {error_msg}", 2)
        return {
            "success": False,
            "error": error_msg
        }

def get_dropdown_options():
    """Return dropdown options for Electron UI"""
    return {
        "locations": [
            "Home",
            "Curzon Building, BCU City Centre",
            "Millenium Point, BCU City Centre",
            "SteamHouse, BCU City Centre",
            "BBC London Broadcasting House",
            "BBC Salford MediaCityUK",
            "BBC Cymru Wales",
            "BBC Scotland, Pacific Quay",
            "BBC Belfast Broadcasting House",
            "BBC Wood Norton",
            "BBC Bristol"
        ],
        "activityTypes": [
            "Annual Leave",
            "Assignment Writing",
            "Combination of activities",
            "External Practice Exposure",
            "Lecture",
            "Other",
            "Placement",
            "Protected Learning",
            "Reading",
            "Research",
            "Revision",
            "Seminar",
            "Tutorial",
            "Work shadowing",
            "Independent Study"
        ],
        "moduleCodes": [
            "Not applicable",
            "DIG4142",
            "CMP4267",
            "ENG4099",
            "CMP4286",
            "DIG4143",
            "ENG4098",
            "ENG5139",
            "CMP5346",
            "CMP5347",
            "CMP5345",
            "CMP5348",
            "DIG5130",
            "DIG6204",
            "CMP6195",
            "DIG6209",
            "DIG6202",
            "DIG6203"
        ],
        "declarationOptions": [
            "Yes",
            "No"
        ],
        "confirmationOptions": [
            "Yes",
            "No",
            "Not applicable"
        ]
    }

def main():
    """Main function - handle command line arguments"""
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No command provided"}))
        sys.exit(1)
    
    command = sys.argv[1]
    
    try:
        if command == "get_options":
            # Return dropdown options
            result = get_dropdown_options()
            print(json.dumps(result))
            
        elif command == "process_data":
            # Process form data
            if len(sys.argv) < 3:
                print(json.dumps({"error": "No data provided"}))
                sys.exit(1)
            
            # Parse JSON data from command line
            json_data = sys.argv[2]
            data = json.loads(json_data)
            
            # Process the data
            result = process_form_data(data)
            print(json.dumps(result))
            
        elif command == "find_next_row":
            # Find the next available row in Excel
            settings = load_settings_from_electron()
            excel_path = settings.get('excelPath', '')
            
            if not excel_path:
                print(json.dumps({"error": "Excel path not set in settings"}))
                sys.exit(1)
                
            if not os.path.exists(excel_path):
                print(json.dumps({"error": f"Excel file not found: {excel_path}"}))
                sys.exit(1)
            
            # Set global path for findFirstBlankRow to use
            global path
            path = excel_path
            
            try:
                next_row = findFirstBlankRow()
                result = {
                    "row": next_row,
                    "starting_row": settings.get('startingRow', 125),
                    "excel_path": excel_path
                }
                print(json.dumps(result))
            except Exception as e:
                print(json.dumps({"error": f"Failed to find next row: {str(e)}"}))
                sys.exit(1)
                
        elif command == "test_connection":
            # Test Python environment
            settings = load_settings_from_electron()
            result = {
                "python_version": sys.version,
                "working_directory": os.getcwd(),
                "settings_loaded": bool(settings),
                "excel_path": settings.get('excelPath', 'Not set')
            }
            print(json.dumps(result))
            
        else:
            print(json.dumps({"error": f"Unknown command: {command}"}))
            sys.exit(1)
            
    except Exception as e:
        error_result = {
            "error": str(e),
            "type": type(e).__name__
        }
        print(json.dumps(error_result))
        sys.exit(1)

if __name__ == "__main__":
    main()