# =============================================================================
## OTJ_AUTOMATION BY LIAM SHADWELL
## COPYTIGHT LIAM SHADWELL ;)
# =============================================================================


## TEXT FILE FUNCTIONS ##

def path():
    global path     #global variable 'path' to be used in other functions for finding the file path
    name = "FilePath.txt"   #set name of File Path config txt file
    path = ""   #set path to blank string
    try: 
        path = readTXT(name)[0].replace("\n","")    #set path as the name specified in the txt file
    except Exception:   #error handline
        fatal(f"In function 'path()' - Cannot find '{name}'")   #this error is fatal, so fatal() is used
    return path     #return path (will be blank due to path = "" if there is exception)


def wrtTXT(file,line):  #function to write txt to a requested txt file.
    filename = open(file, 'a') #open file in append mode
    filename.write(line) # write the specified line to the end of the file
    filename.close() # close file 
    
    
def readTXT(name): #reads text files
    filename = open(name, 'r') #open file in read mode
    contents = filename.readlines() # read all content from the file and store in var
    filename.close() # close file
    return contents #return the data


def findSetting(name):      #function to find a specified setting from settings.txt
    file = "Backend Files (Hidden)/settings.txt"    #specify file
    
    contents = readTXT(file)    #get all data in the above file
    name = f"{name}:"   #add a colon to the name specified when calling this function
    
    for i in range (len(contents)):     #loop for length of txt file
        if (removeNewLines(contents[i])) == (removeNewLines(name)):     #if name of setting asked for is == the setting found during iteraton
            line = i    #set the line number as i to specifiy the line the name of the setting is found on 
    
    output = (contents[line+1]).replace("\n","")    #set output as the line of the setting name +1 (i.e. the setting option)

    return output   #return the output
    

def getInitNotes():     #function to get the init notes from the settings.txt
    received = findSetting("Init Notes")    #uses findSetting() to get the specified init notes
    
    receivedSplit = received.replace("/","\n")      #replace '/' in the txt file with '\n' to format better in CLI
    
    return receivedSplit    #return the notes

#
#
#
#
#
#

## READ EXCEL FUNCTIONS ##

def readCell(row,col): #reads cell with specific coordinate
    wb = load_workbook(path) #open file
    sheet = wb["OTJ log"] #set active sheet 
    
    contents = (sheet.cell(row=row, column=col).value) #get all data from the row and col specified
    contents = str(contents) #convert data to string for processing
    
    if contents == "None": #return data if valid
        return(f"NO DATA IN CELL {col}/{row}")
    else:
        return(contents)
    

def readDate(row):      #specific function to read a date from a specified row
    contents = readCell(row,3)      #uses readCell() to find the data
    if "NO DATA" in contents:   #readCell returns a long string including 'NO DATA'. If present, return the contents as is. 
        return contents
    else:   #if there is data, then remove the time that Excel automatically adds to the cell
        contents = contents[:-9]
    return contents     #return



def findInCol(term,col,style):      #function to find a term in a column and return as either the raw text, or total instances of the term
    wb = load_workbook(path)    #open workbook and sheet
    sheet = wb["OTJ log"]
    
    output = []     #define output list
    total = 0       #define total variable
    
    for iterateRows in range (18, sheet.max_row):       #iterate through all rows
        contents = (sheet.cell(row=iterateRows, column=col).value)      #read the cell contents 
        contents = str(contents)    #convert to str
        
        if term in contents:        #if the term can be found in the cell's contents, then add the Cell's location to the output list
            output.append(f"{iterateRows}/{col}")
            total += 1      #used for style 2 (total instances). Increase total by 1
                
    if output and style == 1:   #return output list for style 1
        return output
    elif output and style == 2: #return total count for style 2
        return total
    else:
        return "Not found"  #if not found, return not found
    

def findFirstBlankRow():    #finds the first blank row in the .xlsx where the data should be written. Note, this is the first row with no data, not the first with no formatting (thats ~2000). 
    return int((findInCol("None",3,1)[0][:-2]))    #uses find in col, passed 'None' and asks for column 3 style 1, which returns the coordinates. 


#
#
#
#
#
#

## WRITING TO EXCEL ##

def writeRow():     #function to write rowData list to excel file. This is run after all the required info from the below functions has been collated. 
    global rowData  #global to be used in the other functions below
    
    print("\nAdding this data to the log. Please wait...")    #status message to user

    wb = load_workbook(path)    #open workbook
    sheet = wb["OTJ log"]   #open sheet
    
    row = findFirstBlankRow()    #get the first blank row available
    
    for col, value in enumerate(rowData, start=3):  # start=1 means column A
        sheet.cell(row=row, column=col, value=value)
    
    wb.save(path)   #save file
    rowData = []    #reset row data list to avoid duplicate entries
    
    print("This data has now been added to the log.")   #print status message
    log(f"writeRow() - Wrote 'rowData' list to '{path}'",1)      #add to log
    

def addDate():  #function to get the required date from the user
    global date     #global to be used in addAcadYear()
    
    date = input("Enter date of activity in format 'DD/MM/YYYY' (or enter 'today'):    ")
    date = date.lower()     #enter required date and format
    
    if date == "today":     #user is able to enter 'today' to get the current date
        date = (datetime.now()).strftime("%d/%m/%Y")    #format date as required
    elif len(date.split("/")) != 3 or len(date.split("/")[2]) != 4 or len(date.split("/")[1]) != 2 or len(date.split("/")[0]) != 2:
        print("Date is not valid. Must be in form DD/MM/YYYY\n")
        addDate()
        return
    
    rowData.append(date)    #add to rowData list

    print(f"Got it, using '{date}'.")   #print status and log entry
    log(f"addDate() - Added date:'{date}' to 'rowData' list",1)


def addAcadYear():  #function to use global data var above and get the current academic year, e.g. 25/26
    try: 
        split = date.split("/")     #take the date given in addDate() and split it into D/M/Y
        
        year = int(split[2][-2:])   #set the year to be the last 2 digits of the year field from the split date, e.g. '26'
        month = int(split[1])       #same for the month
        
        if month >= 9:          #if month is > 9, then the acad year must be the current year
            startYear = year    
        else: 
            startYear = year-1    #else, the start year must be -1 from the current
            
        endYear = startYear + 1     #add 1 to the end year to allow for the acad year after the /
        
        rowData.append(f"{startYear}/{endYear}")    #add the academic year to the rowData list 
        log(f"addAcadYear() - Added academic year {startYear}/{endYear} to rowData list.",1)    #add to the log
    
    except ValueError as e:     #if the wrong data type is entered, then state and log
        print("INCORRECT VALUE TYPE")
        fatal(f"addAcadYear() - Incorrect value type - {e}")
        
    except Exception as e:  #any other exception, kill and log
        fatal(f"addAcadYear() - Fatal error - {e}")

        
    
    
    
def addLocation():  #adding activity location to the rowData list
    try: 
        selected = 0    #define var
        print("\nPlease choose a location: \n")     
        
        options = [     #defining options list, each will have a selectable index number
            "Home",
            "Curzon Building, BCU City Centre",
            "Millenium Point, BCU City Centre",
            "SteamHouse, BCU City Centre",
            "BBC London Broadcasting House",
            "BBC Salford MediaCityUK",
            "BBC Cymru Wales",
            "BBC Scotland, Pacific Quay",
            "BBC Belfast Broadcasting House",
            "BBC Wood Norton",
            "BBC Bristol"
            ]
        
        
        for i in range (1,len(options)):    #define the indexes of the options
            print(f"{i} - {options[i]}")    #and print
            
        selected = int(input("\nEnter the chosen location here >>     "))   #get the users selection
        
        if selected > len(options)-1 or selected == 0:      #validate input, if invalid, then go to start of function
            print("Sorry, that option is not valid. Please try again...\n")
            addLocation()
            
        rowData.append(options[selected])       #add selection to rowData
    
        print(f"Got it, using '{options[selected]}'.")      #status and log
        log(f"addLocation() - Added Location:'{options[selected]}' to 'rowData' list",1)
        
    except ValueError as e:         #if the wrong data type is entered, then state and log
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"addLocation() - {e} - looping to start of function",2)
        addLocation()
        
    except Exception as e:  #any other exception, kill and log
        fatal(f"addLocation() - Fatal error - {e}")
    
    
def addType():   #adding activity type to the rowData list
    try:
        print("\nPlease choose an activity type: \n")
        
        selected = 0    #define var

        options = [        #defining options list, each will have a selectable index number
            "Annual Leave",
            "Assignment Writing",
            "Combination of activities",
            "External Practice Exposure",
            "Lecture",
            "Other",
            "Placement",
            "Protected Learning",
            "Reading",
            "Research",
            "Revision",
            "Seminar",
            "Tutorial",
            "Work shadowing",
            "Independent Study"
        ]

        for i in range (1,len(options)):     #define the indexes of the options
            print(f"{i} - {options[i]}")    #and print
            
        selected = int(input("\nEnter the chosen location here >>     "))   #get the users selection
        
        if selected > len(options)-1 or selected == 0:          #validate input, if invalid, then go to start of function
            print("Sorry, that option is not valid. Please try again...\n")
            addType()
            
        rowData.append(options[selected])   #add selection to rowData
    
        print(f"Got it, using '{options[selected]}'.")          #status and log
        log(f"addType() - Added Type:'{options[selected]}' to 'rowData' list",1)
        
    except ValueError as e:       #if the wrong data type is entered, then state and log
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"addType() - {e} - looping to start of function",2)
        addType()
        
    except Exception as e:     #any other exception, kill and log
        fatal(f"addType() - Fatal error - {e}")
   
        
def addModule():      #see above
    try:
        print("\nPlease choose a module code: \n")
        
        selected = 0

        options = [
            "Not applicable",
            "DIG4142",
            "CMP4267",
            "ENG4099",
            "CMP4286",
            "DIG4143",
            "ENG4098",
            "ENG5139",
            "CMP5346",
            "CMP5347",
            "CMP5345",
            "CMP5348",
            "DIG5130",
            "DIG6204",
            "CMP6195",
            "DIG6209",
            "DIG6202",
            "DIG6203"
        ]

        for i in range (len(options)):
            print(f"{i+1} - {options[i]}")
            
        selected = int(input("\nEnter the chosen module here >>     "))
        
        if selected < 1 or selected > len(options): 
            print("Sorry, that option is not valid. Please try again...\n")
            addType()
            return #ensure full exit post recursion
            
        rowData.append(options[selected-1])
    
        print(f"Got it, using '{options[selected-1]}'.")    
        log(f"addModule() - Added Module: {options[selected-1]}' to 'rowData' list",1)
        
    except ValueError as e:
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"addModule() - {e} - looping to start of function",2)
        addType()
        
    except Exception as e:
        fatal(f"addModule() - Fatal error - {e}")
        
        
def addDescription():       #function to add description of activity to rowData
    try:    
        line = ""   #define var
        line = input("\nEnter a breif overview of what you did in this session....\n\n")    #take input
        
        if line != "":  #if input is valid
            rowData.append(line)    #add input to rowData
            print("\n\nCool, we'll add that to the OTJ log.")   #print and log
            log("addDescription() - Added description to 'rowData' list.",1)
        else:   #if invalid
            print("\n\nSorry, that's invalid...")
            log(f"addDescription() - Invalid input '{line}'",2)
            addDescription()    #go to start of function
            
    except Exception as e:      #if fatal error, use fatal()
        fatal(f"addDescription() - Fatal error - {e}")


def addDetails():       #function to add details of activity to rowData
    try:
        line = input("\nEnter more details about what you did in this session....\n\n")     #take input
        
        if line != "":      #if valid, add to rowData
            rowData.append(line)    
            print("\n\nCool, we'll add that to the OTJ log.")
        else:   #else, restart function
            print("\n\nSorry, this is a required field...")
            log(f"addDetails() - Invalid input - '{line}'")
            addDetails()    
             
    except Exception as e:      #error handling
        fatal(f"addDetails() - Fatal error - {e}")
        
        
def addKSB():       #add KSBs to rowData
    try: 
        if rowData[4] != "Not applicable":      #only if module code provided
            KSB = getKSB(rowData[4])    #use getKSB() to get the KSBs for the provided code
            if KSB:
                print("KSBs received")
            else:
                fatal("addKSB() - Unable to get KSBs for module code provided")
            rowData.append(KSB)     #add to rowData
    except Exception as e:  #error handling
        fatal(f"addKSB() - Unable to get KSBs for module code provided - {e}")
        
        
def addNextSteps():     #add next steps to rowData
    try: 
        line = ""      #define var
        line = input("\nEnter any next steps if applicable....\n\n")    #take input
        
        rowData.append(line)    #add to rowData
        
        if line =="":       #Next Steps is optional data, so print correct line depending on if input provided
            print("No worries, we won't add anything\n")
        else: 
            print("Cool, we'll add that to the OTJ log.")
    
    except Exception as e:  #handle fatal errors
        fatal(f"addNextSteps() - Fatal error - {e}")
    

def addDuration():      #add activity duration to rowData
    try: 
        hours = int(input("Enter the hours spent during this session (e.g. 2.5)...     "))  #take input
    
        if hours <= 0 or hours >= 50:   #if input invalid
            print("Sorry, that's not valid. The input must be between 0 and 50 hours. Try again\n")
            addDuration()   #go to start of function
        else: 
            rowData.append(hours)   #if valid, add to rowData
            
    except ValueError as e:     #handle value errors and restart function
        print("INCORRECT VALUE TYPE ENTERED\nTry again\n\n")
        log(f"addDuration() - ValueError - {e}")
        addDuration()
        
    except Exception as e:  #handle fatal errors
        fatal(f"addDuration() - Fatal error - {e}")
        
        
def addDeclaration():   #used to edit column M in the .xlsx, to confirm completed in normal working hours
    try:
        print("\nHas this been completed within normal working hours?\n")
        
        selected = 0    

        options = [
            "Yes",
            "No"
        ]

        for i in range (len(options)):
            print(f"{i+1} - {options[i]}")
            
        selected = int(input("\nEnter selection >>     "))
        
        if selected < 1 or selected > len(options): 
            print("Sorry, that option is not valid. Please try again...\n")
            addDeclaration()
            return
            
        rowData.append(options[selected-1])
    
        print(f"Got it, using '{options[selected-1]}'.")
        log(f"addDeclaration() - Added Declaration: {options[selected-1]}' to 'rowData' list",1)
        
        if options[selected-1] == "No":
            addConfirmation()
        else:
            rowData.append("Not applicable")
        
    except ValueError as e:
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"addDeclaration() - {e} - looping to start of function",2)
        addType()
        
    except Exception as e:
        fatal(f"addDeclaration() - Fatal error - {e}")
        

def addConfirmation():
    try:
        print("\nHas this been approved by the employer?\n")
        
        selected = 0

        options = [
            "Yes",
            "No",
            "Not applicable"
        ]

        for i in range (len(options)):
            print(f"{i+1} - {options[i]}")
            
        selected = int(input("\nEnter selection >>     "))
        
        if selected < 1 or selected > len(options): 
            print("Sorry, that option is not valid. Please try again...\n")
            addConfirmation()
            return
            
        rowData.append(options[selected-1])
    
        print(f"Got it, using '{options[selected-1]}'.")
        log(f"addConfirmation() - Added Confirmation: {options[selected-1]}' to 'rowData' list",1)
        
    except ValueError as e:
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"addConfirmation() - {e} - looping to start of function",2)
        addType()
        
    except Exception as e:
        fatal(f"addConfirmation() - Fatal error - {e}")
        
        


#
#
#
#
#
#


## READ EXCEL FOR KSBs ##

def readCellKSB(row,col): #as above, but specific to KSB sheet
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]#KSB data sheet
    
    contents = (sheet.cell(row=row, column=col).value)
    contents = str(contents)
    
    if contents == "None":
        return(f"NO DATA IN CELL {col}/{row}")
    else:
        return(contents)
    
    
def findInRowKSB(term,style):   #as above, but using sheet Broadcast & Media KSBs
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]    
    
    output = []
    total = 0
    
    for iterateCols in range (3, sheet.max_column):
        contents = (sheet.cell(row=2, column=iterateCols).value)
        contents = str(contents)
        
        if term in contents: 
            output.append(f"2/{iterateCols}")
            total += 1
            
    if output and style == 1: 
        return output
    elif output and style == 2: 
        return total
    else:
        return "Not found"
    
    
def findInColKSB(term,col,style):       #as seen in findInCol(), but using the sheet with the KSBs defined
    wb = load_workbook(path)
    sheet = wb["Broadcast & Media KSBs"]    
    
    output = []
    total = 0
    
    for iterateRows in range (1, sheet.max_row):
        contents = (sheet.cell(row=iterateRows, column=col).value)
        contents = str(contents)
        
        if term in contents: 
            output.append(f"{iterateRows}/{col}")
            total += 1
                
    if output and style == 1: 
        return output
    elif output and style == 2: 
        return total
    else:
        return "Not found"
    
    
def getKSB(module):  #gets KSBs given a module
    
    module = module.upper()     #input validation
    
    print(f"\nGetting KSBs for module {module} - please wait...")    #waiting message
    
    rowsWithX = []      #init lists
    fullKSB = []
    
    received = findInRowKSB(module,1)   #find in KSB sheet header row the column which matches the required module code. 
    
    column = (str(received[0]))     #string
    column = column.split("/")  #split by /
    column = int(column[1])     #set the column number as int for later use. 
    
    rowsWithXOutput = findInColKSB("x", column, 1)  #check any rows in the column found above for 'x', denoting they are linked to the inputted module. 

    for item in rowsWithXOutput:    
        parts = item.split("/")   # Split the string at '/'
        number_after_slash = parts[0]   # Take the part after the slash
        rowsWithX.append(number_after_slash)  # Add it to the new list
            
    for item in rowsWithX:
        fullKSB.append(readCellKSB(int(item),2)[:2])
        
    return str(",".join(fullKSB))


#
#
#
#
#
#
    
## MISC FUNCTIONS ## 

def t():    #subroutine to easily print a test line
    print("test")
    
    
def removeNewLines(string):     #function to remove new lines where needed, instead of typing out '.replace(fubfj)
    try:    #error handling
        output = string.replace("\n","")    #take the inputted string, and remove the \n
        return output
    except Exception as e:   #if error, then file name exception could occur, requiring fatal error closure
        fatal(f"In removeNewLines() - Unable to remove new line - {e}")
    

def log(line,style):    #function to log input with either info, error or fatal tags at the start. 
    
        
    currentTime = (datetime.now()).strftime("%Y-%m-%d %H:%M:%S")    #get current time
    
    if style == 1:      #depending on the style, use wrtTXT() to write the output defined to the log.txt file in Backend Files. 
        output = "INFO    - {} - {}\n".format(currentTime,line)
        wrtTXT("Backend Files (Hidden)/log.txt",output)
    elif style == 2:
        output = "#ERROR# - {} - {}\n".format(currentTime,line)
        wrtTXT("Backend Files (Hidden)/log.txt",output)
    elif style == 999:
        output = "#FATAL# - {} - {}\n".format(currentTime,line)
        wrtTXT("Backend Files (Hidden)/log.txt",output)
    
    elif style == "Start":      #used to generate a program started message to clearly show a new instance. 
        output = "\n\n\nPROGRAM STARTED @ {}\n\n".format(currentTime)
        wrtTXT("Backend Files (Hidden)/log.txt",output)
    
    else:   #if the log style cannot be defined, recur back through and send error message.  
        log(f"In function 'log()' - UNKNOWN LOG STYLE '{style}'",2)
        
def fatal(line):    #function to kill program if fatal error
                    #call this function from other functions where a fatal error exception could occur
    log(line,999)   #auto log fatal error
    
    print("\n\n\n")     #print error into CLI
    print("#########################################################################")
    print("FATAL ERROR - PROGRAM FORCE QUIT IN 5 SECONDS - CONTACT LIAM - SEE LOGS")
    print("#########################################################################")
    time.sleep(5)   #wait 5 sec before closing program
    os._exit(1)  #force close 
    
    
def moreEntries():  #ask the user if they want to add more data
    try:  
        print("\nDo you want to add more entries to your OTJ log?\n")   #print line
        
        options = [     #valid options
            "Yes",
            "No"
            ]
        
        for i in range (len(options)):
            print(f"{i+1} - {options[i]}")
            
        choice = int(input("\nEnter selection:    "))     #enter selection
        
        if choice < 1 or choice > len(options):
            print("Sorry, that's an invalid option...")
            moreEntries()
            return
        
        else: 
            selection = options[choice-1]
        
        if selection == 'Yes':
            runProgram()
            log("moreEntries() - looping to start of program in 'runProgram()'",1)
        elif selection == "No":
            input("No worries, hit enter again to quit the program...     ")    
        else:
            print("That's an invalid selection")
            log("moreEntries() - Invalid selection - looping to start of function",2)
            moreEntries()
            return
            
    except ValueError as e:
        print("INCORRECT VARIABLE TYPE ENTERED! Try again...")
        log(f"moreEntries() - {e} - looping to start of function",2)
        moreEntries()
        return
            
    except Exception as e:   #fatal code if req
        fatal(f"moreEntries() - {e}")
        
        
        
def runProgram():
    try:
        addDate()
        addAcadYear()
        addLocation()
        addType()
        addModule()
        addDescription()
        addDetails()
        addKSB()
        addNextSteps()
        addDuration()
        addDeclaration()
        
        writeRow()
    
        moreEntries()
        
    except Exception as e:   #fatal code if req
        fatal(f"runProgram() - {e}")
        
    
        
#
#
#
#
#
#

## MISC OTHER ##

import warnings    #import warnings to silence any data warnings from OpenPyXL in the running CLI
warnings.simplefilter("ignore", UserWarning)

from openpyxl import load_workbook   #import openpyxl
from datetime import datetime    #these are pretty obvious
import time
import os

# Get the directory of the current script
current_dir = os.path.dirname(os.path.abspath(__file__))
# Change working directory to that folder
os.chdir(current_dir)

rowData = []    #define for adding entries to instances of writing rows


#
#
#
#
#
#

## MAIN PROGRAM ##

log("","Start")   #adds a starting line to the log for the current instance

path()  #gets the file path for the Excel file

if path != "":  #if path exists
    print(getInitNotes())
    runProgram()
    


